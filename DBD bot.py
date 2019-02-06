import threading
import discord
import asyncio
from itertools import cycle
import random
import openpyxl
import os



import datetime
from discord.ext import commands

client = discord.Client()
Client = commands.Bot(command_prefix = ';')


status = ['Dead By Daylight Bot!', "I'm pengmin", 'Enjoy~']


def TaskB():
    print('Process B')
    threading.Timer(600, TaskB).start()

TaskB()

@client.event
async def on_member_join(member):
    role = ""
    for i in member.server.roles:
        if i.name == "새로 오신 분!":
            role = i
            break
    await client.add_roles(member, role)

@client.event
async def on_ready():
    print("login")
    print(client.user.name)
    print(client.user.id)
    print("======================")

async def change_status():
    await client.wait_until_ready()
    msgs = cycle(status)

    while not client.is_closed:
        current_status = next(msgs)
        await client.change_presence(game = discord.Game(name = current_status))
        print("봇아 꺼지지 마라...")
        await asyncio.sleep(10)


@client.event
async def on_message(message):
    if message.content.startswith("!카즈"):
        await client.send_message(message.channel,"카즈 변태래요~")
    """
    if message.content.startswith("!수지"):
        await client.send_message(message.channel, "https://bit.ly/2G87EWI")
        await client.send_message(message.channel, "여러분~ 모두 태보하세요~")
        """

    if message.content.startswith("!명령어"):
        channel = message.channel
        embed = discord.Embed(
            title='꿀바데 봇의 명령어 입니다.',
            description='',
            colour=discord.Colour.blue()
        )


        embed.add_field(name=':heart: !출첵', value=': 매일매일 24시간에 한 번 출석체크가 가능합니다.', inline=False)
        embed.add_field(name=':heart: !출확', value=': 몇 번 출석했는지 확인 할 수 있습니다.', inline=False)
        embed.add_field(name=':heart: !살인마정하기 사람1 사람2 사람3 사람4 사람5', value=": 다섯 사람 중 살인마를 뽑을 수 있습니다.", inline=False)
        embed.add_field(name=':heart: !프로필 스팀이름', value=": 스팀에서 프로필을 검색할 수 있습니다.", inline=False)
        embed.add_field(name = ':heart: !건의 건의내용', value =': 추가됐으면 좋겠다는 기능을 건의할 수 있습니다.', inline = False)

        await client.send_message(channel, embed=embed)

    if message.content.startswith("!건의"):
        multi = message.content.split(" ")
        contents = multi[1:]
        contents = str(contents)
        member = message.author.name
        file = openpyxl.load_workbook("건의사항.xlsx")
        sheet = file.active
        for i in range(1, 101):
            if str(sheet["A" + str(i)].value) == '-':
                sheet["A" + str(i)].value = contents
                sheet["B" + str(i)].value = member
                await client.send_message(message.channel, "건의사항이 저장되었습니다.")
                file.save("건의사항.xlsx")
                break
            elif str(sheet["A" + str(i)].value) != '-':
                file.save("건의사항.xlsx")
                break

    if message.content.startswith("!출첵"):
        file = openpyxl.load_workbook("쿨타임(쳌).xlsx")
        sheet = file.active
        for i in range(1, 298):
            if str(sheet["A" + str(i)].value) == str(message.author.id):
                if int(sheet["B" + str(i)].value) <= int(datetime.datetime.now().strftime("%Y%m%d%H%M%S")):
                    await client.send_message(message.channel, "출석체크 완료!")
                    a = datetime.datetime.now() + datetime.timedelta(hours = 24)
                    sheet["B" + str(i)].value = a.strftime("%Y%m%d%H%M%S")

                    sheet["C" + str(i)].value += 1




                    file.save("쿨타임(쳌).xlsx")

                    break
                else:
                    checktime = sheet["B" + str(i)].value
                    checktime = str(checktime)
                    year = int(checktime[0:4])
                    month = int(checktime[4:6])
                    day = int(checktime[6:8])
                    hour = int(checktime[8:10])
                    minute = int(checktime[10:12])
                    second = int(checktime[12:14])
                    plan_time = datetime.datetime(year, month, day, hour, minute, second)
                    print(plan_time)
                    tdy = datetime.datetime.now()
                    time_remainer = plan_time - tdy
                    time_remainer = time_remainer.seconds
                    time_remainer_hour = int(time_remainer / 3600)
                    time_remainer_minute = int((time_remainer - (time_remainer_hour * 3600)) / 60)
                    time_remainer_second = int(time_remainer - (time_remainer_hour * 3600) - (time_remainer_minute * 60))

                    await client.send_message(message.channel, "남은 시간 : {}시간 {}분 {}초. 예상날짜 : {}일 {}시 {}분 {}초".format(time_remainer_hour, time_remainer_minute, time_remainer_second, day, hour, minute, second))
                    break
            if str(sheet["A" + str(i)].value) == '-':
                sheet["A" + str(i)].value = str(message.author.id)
                a = datetime.datetime.now() + datetime.timedelta(hours = 24)
                sheet["B" + str(i)].value = a.strftime("%Y%m%d%H%M%S")

                sheet["C" + str(i)].value += 1

                file.save("쿨타임(쳌).xlsx")

                await client.send_message(message.channel, "출석체크 완료!")
                break

    if message.content.startswith("!출확"):
        file = openpyxl.load_workbook("쿨타임(쳌).xlsx")

        many = message.content.split(" ")
        name = many[1:]  # 사람 이름
        length = len(message.author.roles)
        role = message.author.roles[length - 1].name
        sheet = file.active

        for i in range(1, 298):
            if str(sheet["A" + str(i)].value) == message.author.id:
                times = str(sheet["C" + str(i)].value)
                var_time = int(times)
                length = len(message.author.roles)



                if sheet["C" + str(i)].value == 1:
                    """
                    role = ""
                    member = discord.utils.get(client.get_all_members(), id=str(message.author.id))
                    for i in member.server.roles:
                        if i.name == ":blue_heart: 청정수":
                            role = i
                            break

                    await client.add_roles(member, role)
                    """
                    await client.send_message(message.channel, "!등급 ")
                    if message.content.startswith("!등급"):  # 다른 사람들도 명령어를 사용 가능한 심각한 버그에 걸림
                        role = ""
                        rolename = message.content.split(" ")
                        member = discord.utils.get(client.get_all_members(), id=str(rolename[1]))
                        roletrue = rolename[2:]
                        name = ""
                        for i in roletrue:
                            name = name + i + " "
                        name = name[:-1]
                        for i in message.server.roles:
                            if i.name == name:
                                role = i
                                break
                        await client.add_roles(member, role)

                embed = discord.Embed(
                    title= 'Profile'
                           ,
                    colour=discord.Colour.orange()
                )
                embed.add_field(name=' 플레이어 정보', value= message.author, inline=False)
                embed.add_field(name=' 플레이어 등급', value= role, inline=False)
                embed.add_field(name=' 출석횟수', value='지금까지 ' + times + '번 출석하셨습니다.', inline=False)

                await client.send_message(message.channel, embed=embed)
                file.save("쿨타임(쳌).xlsx")
                break
    if message.content.startswith("!살인마정하기"):
        person = message.content.split(" ")
        person = person[1:]
        random.shuffle(person)
        await client.send_message(message.channel, person[1] + "→" + '살인마')

    if message.content.startswith("!프로필"):
        content = message.content.split(" ")
        id = content[1]
        await client.send_message(message.channel, "https://steamcommunity.com/search/users/#text=" + id)



    if message.content.startswith("!등급"): # 다른 사람들도 명령어를 사용 가능한 심각한 버그에 걸림

        role = ""
        rolename = message.content.split(" ")
        member = discord.utils.get(client.get_all_members(), id = str(rolename[1]))
        roletrue = rolename[2:]
        name = ""
        for i in roletrue:
            name = name + i + " "
        name = name[:-1]
        for i in message.server.roles:
            if i.name == name:
                role = i
                break
        await client.add_roles(member, role)

    if message.content.startswith("!등삭"): # 다른 사람들도 명령어를 사용 가능한 심각한 버그에 걸림
        role = ""
        rolename = message.content.split(" ")
        member = discord.utils.get(client.get_all_members(), id=str(rolename[1]))

        roletrue = rolename[2:]
        name = ""
        for i in roletrue:
            name = name + i + " "
        name = name[:-1]
        for i in message.server.roles:
            if i.name == name:
                role = i
                break
        await client.remove_roles(member, role)

    if message.content.startswith(";clear"):
        con = message.content.split(" ")
        ctx = con[1]
        amount = 100
        channel = message.channel
        messages = []
        async for message in client.logs_from(channel, limit = int(ctx)):
            messages.append(message)
        await client.delete_messages(messages)
"""
    if message.content.startswith("!출석보정"):
        mes = message.content.split(" ")
        id = mes[1]
        count = int(mes[2])
        file = openpyxl.load_workbook("쿨타임(쳌).xlsx")
        sheet = file.active
        for i in range(1, 101):
            if str(sheet["A" + str(i)].value) == str(id):

                sheet["C" + str(i)].value += count
                file.save("쿨타임(쳌).xlsx")
                break
            if str(sheet["A" + str(i)].value) == '-':
                sheet["A" + str(i)].value = str(id)
                sheet["C" + str(i)].value += count

                file.save("쿨타임(쳌).xlsx")

                await client.send_message(message.channel, "출석체크 완료!")
                break
                """





access_token = os.environ["BOT_TOKEN"]
client.run(access_token)
